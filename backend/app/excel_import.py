"""Safe Excel import helpers for the Evaz Property Index admin panel.

The parser intentionally validates before writing anything.  It reads the
header row (row 3 in the official workbook), preserves unknown columns for
diagnostics, and returns row-level errors that the UI can show to the operator.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


SHEETS = {
    "ورود ملک": "properties",
    "مناطق و محله‌ها": "neighborhoods",
    "رکوردهای بازار": "market_records",
    "معاملات قطعی": "transactions",
    "آگهی‌ها": "listings",
    "ارزیابی قیمت": "valuations",
    "مسکن مهر": "mehr_housing",
    "مسکن ملی": "national_housing",
    "منابع داده": "data_sources",
    "فرصت ویژه ماه": "special_opportunities",
    "رتبه کارشناسی مناطق": "region_rankings",
}

REQUIRED = {
    "ورود ملک": ["کد داخلی ملک", "نوع ملک"],
    "مناطق و محله‌ها": ["کد منطقه", "نام منطقه", "کد محله", "نام محله"],
    "رکوردهای بازار": ["کد رکورد بازار", "کد داخلی ملک", "نوع رکورد", "تاریخ", "قیمت کل تومان"],
    "معاملات قطعی": ["کد رکورد بازار", "کد داخلی ملک", "تاریخ قرارداد", "قیمت نهایی تومان"],
    "آگهی‌ها": ["کد آگهی", "کد رکورد بازار", "کد داخلی ملک", "قیمت فروشنده تومان"],
    "ارزیابی قیمت": ["کد ارزیابی", "کد رکورد بازار", "کد داخلی ملک", "تاریخ ارزیابی"],
    "مسکن مهر": ["کد داخلی ملک", "حد پایین تومان", "حد بالا تومان"],
    "مسکن ملی": ["کد ملک", "فاز", "حد پایین تومان", "حد بالا تومان"],
    "منابع داده": ["کد منبع", "نام منبع"],
    "فرصت ویژه ماه": ["کد فرصت ویژه", "ماه", "کد آگهی", "کد داخلی ملک"],
    "رتبه کارشناسی مناطق": ["دوره", "منطقه", "رتبه ۱ تا ۱۰"],
}


def _clean(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip() or None
    return value


def _number(value: Any, field: str, errors: list[str], positive: bool = False) -> int | float | None:
    value = _clean(value)
    if value is None:
        return None
    if isinstance(value, bool):
        errors.append(f"{field}: مقدار عددی نیست")
        return None
    try:
        number = float(str(value).replace(",", "").replace("٬", ""))
        if positive and number <= 0:
            raise ValueError
        return int(number) if number.is_integer() else number
    except (TypeError, ValueError):
        errors.append(f"{field}: مقدار عددی نامعتبر")
        return None


def parse_workbook(content: bytes) -> dict[str, Any]:
    wb = load_workbook(BytesIO(content), data_only=False, read_only=True)
    result: dict[str, Any] = {"sheets": [], "total_rows": 0, "total_errors": 0}
    for ws in wb.worksheets:
        if ws.title not in SHEETS or ws.title in {"راهنما", "فهرست‌های مجاز"}:
            continue
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            result["sheets"].append({"name": ws.title, "status": "error", "errors": ["ردیف سرستون پیدا نشد"]})
            result["total_errors"] += 1
            continue
        headers = [_clean(v) for v in rows[2]]
        headers = [str(v) if v is not None else "" for v in headers]
        missing = [h for h in REQUIRED.get(ws.title, []) if h not in headers]
        sheet_result = {"name": ws.title, "entity": SHEETS[ws.title], "rows": [], "row_count": 0, "error_count": 0, "errors": []}
        if missing:
            sheet_result["errors"].append("ستون‌های الزامی پیدا نشد: " + "، ".join(missing))
        for row_number, raw in enumerate(rows[3:], start=4):
            values = [_clean(v) for v in raw]
            if not any(v is not None for v in values):
                continue
            # The official template pre-fills QA formulas far below the sample
            # area. They are not data rows and must not be imported.
            non_qa = values[:-1] if values else values
            if not any(v is not None for v in non_qa) and isinstance(values[-1] if values else None, str) and str(values[-1]).startswith("="):
                continue
            record = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
            errors: list[str] = []
            for required in REQUIRED.get(ws.title, []):
                if record.get(required) in (None, ""):
                    errors.append(f"{required}: الزامی است")
            for field in ("قیمت کل تومان", "قیمت نهایی تومان", "قیمت فروشنده تومان", "حد پایین تومان", "حد بالا تومان", "پرداخت‌شده تومان", "تعهد باقی‌مانده تومان", "قیمت بازار/انتقال تومان"):
                if field in record and record[field] is not None:
                    _number(record[field], field, errors, positive=True)
            if "درصد پیشرفت" in record and record["درصد پیشرفت"] is not None:
                progress = _number(record["درصد پیشرفت"], "درصد پیشرفت", errors)
                if isinstance(progress, (int, float)) and not 0 <= progress <= 100:
                    errors.append("درصد پیشرفت: باید بین ۰ تا ۱۰۰ باشد")
            if "تاریخ" in record and isinstance(record["تاریخ"], str) and record["تاریخ"] > date.today().isoformat():
                errors.append("تاریخ: تاریخ آینده مجاز نیست")
            item = {"row": row_number, "values": record, "errors": errors, "status": "error" if errors else "valid"}
            sheet_result["rows"].append(item)
            sheet_result["row_count"] += 1
            sheet_result["error_count"] += len(errors)
        result["total_rows"] += sheet_result["row_count"]
        result["total_errors"] += sheet_result["error_count"] + len(sheet_result["errors"])
        result["sheets"].append(sheet_result)
    return result


def _date(value: Any) -> str | None:
    value = _clean(value)
    if value is None:
        return None
    text = str(value)
    return text[:10] if len(text) >= 10 and text[4:5] == "-" else None


def commit_workbook(content: bytes, conn, user_id: str | None = None) -> dict[str, Any]:
    """Validate and insert the mappable core sheets atomically.

    The official workbook contains presentation/manual sheets whose schema
    intentionally has no date or foreign-key fields. Those sheets are reported
    as skipped; core property, source, market-record, transaction, listing and
    valuation rows are committed in one transaction.
    """
    report = parse_workbook(content)
    if report["total_errors"]:
        raise ValueError(report)
    by_name = {s["name"]: s for s in report["sheets"]}
    inserted = {"properties": 0, "data_sources": 0, "market_records": 0, "transactions": 0, "listings": 0, "valuations": 0, "mehr_housing": 0, "national_housing": 0, "special_opportunities": 0, "region_rankings": 0}
    skipped: list[str] = []
    with conn.transaction():
        source_ids: dict[str, str] = {}
        for item in by_name.get("منابع داده", {}).get("rows", []):
            v = item["values"]; code = str(v.get("کد منبع") or ""); name = v.get("نام منبع")
            if not name: continue
            found = conn.execute("SELECT id FROM app.data_sources WHERE name=%s ORDER BY id LIMIT 1", (name,)).fetchone()
            if found: source_ids[code] = found["id"]; continue
            row = conn.execute("INSERT INTO app.data_sources(name,source_type,reliability) VALUES (%s,%s,%s) RETURNING id", (name, "excel", v.get("سطح اعتبار"))).fetchone()
            source_ids[code] = row["id"]; inserted["data_sources"] += 1

        property_ids: dict[str, str] = {}
        for item in by_name.get("ورود ملک", {}).get("rows", []):
            v = item["values"]; code = str(v.get("کد داخلی ملک") or ""); type_name = v.get("نوع ملک")
            type_row = conn.execute("SELECT id FROM app.property_types WHERE name=%s OR code=%s LIMIT 1", (type_name, type_name)).fetchone()
            if not type_row: raise ValueError({"row": item["row"], "error": f"نوع ملک پیدا نشد: {type_name}"})
            existing = conn.execute("SELECT id FROM app.properties WHERE public_code=%s", (code,)).fetchone()
            if existing: property_ids[code] = existing["id"]; continue
            region_row = conn.execute("SELECT id FROM app.regions WHERE name=%s OR slug=%s LIMIT 1", (v.get("نام منطقه"), v.get("کد منطقه"))).fetchone()
            neighborhood_row = conn.execute("SELECT id FROM app.neighborhoods WHERE name=%s OR code=%s LIMIT 1", (v.get("نام محله"), v.get("کد محله"))).fetchone()
            row = conn.execute("""INSERT INTO app.properties(
                public_code,property_type_id,area_m2,building_area_m2,floor,build_year,private_address,
                region_id,neighborhood_id,general_area,commercial_area_m2,building_age,floor_count,
                street_width,usage_type,building_quality,building_condition,features,public_notes,private_notes)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (code, type_row["id"], v.get("زمین مترمربع"), v.get("زیربنا مترمربع"), v.get("طبقه"),
                 v.get("سال ساخت"), v.get("آدرس خصوصی"),
                 region_row["id"] if region_row else None, neighborhood_row["id"] if neighborhood_row else None,
                 v.get("محدوده عمومی"), v.get("تجاری مترمربع"), v.get("سن بنا"), v.get("تعداد طبقات"),
                 v.get("عرض گذر"), v.get("کاربری"), v.get("کیفیت بنا"), v.get("وضعیت بنا"),
                 v.get("امکانات"), v.get("توضیحات عمومی"), v.get("یادداشت خصوصی"))).fetchone()
            property_ids[code] = row["id"]; inserted["properties"] += 1

        record_ids: dict[str, str] = {}
        for item in by_name.get("رکوردهای بازار", {}).get("rows", []):
            v = item["values"]; code = str(v.get("کد رکورد بازار") or ""); property_id = property_ids.get(str(v.get("کد داخلی ملک") or ""))
            type_row = conn.execute("SELECT property_type_id AS id FROM app.properties WHERE id=%s", (property_id,)).fetchone() if property_id else None
            region = conn.execute("SELECT id,name FROM app.regions WHERE name=%s OR slug=%s LIMIT 1", (v.get("نام منطقه"), v.get("کد منطقه"))).fetchone()
            if not region: region = conn.execute("SELECT id,name FROM app.regions ORDER BY created_at LIMIT 1").fetchone()
            if not type_row or not region or not _date(v.get("تاریخ")): raise ValueError({"row": item["row"], "error": "نوع ملک، منطقه یا تاریخ رکورد معتبر نیست"})
            row = conn.execute("""INSERT INTO app.market_records(property_id,record_month,property_type_id,region_id,region_name_snapshot,source_id,status,is_public)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""", (property_id, _date(v.get("تاریخ")), type_row["id"], region["id"], region["name"], source_ids.get(str(v.get("کد منبع") or "")), v.get("وضعیت بررسی") or "raw", str(v.get("انتشار عمومی") or "خیر") in ("بله", "فعال", "منتشرشده"))).fetchone()
            record_ids[code] = row["id"]; inserted["market_records"] += 1

        for item in by_name.get("معاملات قطعی", {}).get("rows", []):
            v = item["values"]; record_id = record_ids.get(str(v.get("کد رکورد بازار") or ""))
            if not record_id: raise ValueError({"row": item["row"], "error": "کد رکورد بازار معامله پیدا نشد"})
            conn.execute("""INSERT INTO app.transactions(market_record_id,transaction_month,final_price_toman,is_public)
                VALUES (%s,%s,%s,%s) ON CONFLICT (market_record_id) DO NOTHING""", (record_id, _date(v.get("تاریخ قرارداد")), int(float(v.get("قیمت نهایی تومان"))), str(v.get("نمایش عمومی") or "خیر") == "بله"))
            inserted["transactions"] += 1

        for item in by_name.get("آگهی‌ها", {}).get("rows", []):
            v = item["values"]; record_id = record_ids.get(str(v.get("کد رکورد بازار") or ""))
            if not record_id: raise ValueError({"row": item["row"], "error": "کد رکورد بازار آگهی پیدا نشد"})
            conn.execute("""INSERT INTO app.listings(market_record_id,asking_price_toman,price_per_m2_toman,status,is_public)
                VALUES (%s,%s,%s,%s,%s) ON CONFLICT (market_record_id) DO NOTHING""", (record_id, int(float(v.get("قیمت فروشنده تومان"))), int(float(v.get("قیمت هر متر اختیاری"))) if v.get("قیمت هر متر اختیاری") else None, v.get("وضعیت فایل") or "draft", False))
            inserted["listings"] += 1

        for item in by_name.get("ارزیابی قیمت", {}).get("rows", []):
            v = item["values"]; property_id = property_ids.get(str(v.get("کد داخلی ملک") or ""))
            if not property_id or not _date(v.get("تاریخ ارزیابی")): raise ValueError({"row": item["row"], "error": "ملک یا تاریخ ارزیابی پیدا نشد"})
            conn.execute("""INSERT INTO app.expert_valuations(property_id,valuation_month,estimated_low_toman,estimated_high_toman,method,status)
                VALUES (%s,%s,%s,%s,%s,%s)""", (property_id, _date(v.get("تاریخ ارزیابی")), int(float(v.get("حد پایین تومان"))), int(float(v.get("حد بالا تومان"))), v.get("منبع/ارزیاب") or "excel", "draft"))
            inserted["valuations"] += 1

        for item in by_name.get("مسکن مهر", {}).get("rows", []):
            v=item["values"]; property_id=property_ids.get(str(v.get("کد داخلی ملک") or ""))
            if not property_id: raise ValueError({"row": item["row"], "error": "ملک مسکن مهر ابتدا باید در شیت ورود ملک ثبت شود"})
            conn.execute("INSERT INTO app.mehr_housing_entries(property_id,floor,low_price_toman,high_price_toman,renovation_status,block,delivery_status,features,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)", (property_id,v.get("طبقه"),v.get("حد پایین تومان"),v.get("حد بالا تومان"),v.get("وضعیت بازسازی"),v.get("بلوک"),v.get("تحویل"),v.get("امکانات"),v.get("توضیحات")))
            inserted["mehr_housing"] += 1

        for item in by_name.get("مسکن ملی", {}).get("rows", []):
            v=item["values"]; property_id=property_ids.get(str(v.get("کد ملک") or ""))
            if not property_id: raise ValueError({"row": item["row"], "error": "ملک مسکن ملی ابتدا باید در شیت ورود ملک ثبت شود"})
            conn.execute("INSERT INTO app.national_housing_entries(property_id,phase,progress_stage,low_price_toman,high_price_toman,paid_toman,remaining_commitment_toman,transfer_price_toman,progress_percent,delivery_status,features,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (property_id,v.get("فاز"),v.get("مرحله پیشرفت"),v.get("حد پایین تومان"),v.get("حد بالا تومان"),v.get("پرداخت‌شده تومان"),v.get("تعهد باقی‌مانده تومان"),v.get("قیمت بازار/انتقال تومان"),v.get("درصد پیشرفت"),v.get("تحویل"),v.get("امکانات"),v.get("توضیحات")))
            inserted["national_housing"] += 1

        for item in by_name.get("فرصت ویژه ماه", {}).get("rows", []):
            v=item["values"]
            conn.execute("INSERT INTO app.special_opportunities_entries(code,month,listing_code,property_id,title,price_toman,comparison_percent,reason,advantages,limitations,image_path,publish_date,end_date,contact_text,status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (v.get("کد فرصت ویژه"),v.get("ماه"),v.get("کد آگهی"),property_ids.get(str(v.get("کد داخلی ملک") or "")),v.get("عنوان"),v.get("قیمت تومان"),v.get("درصد مقایسه با میانگین"),v.get("دلیل انتخاب"),v.get("مزایا"),v.get("محدودیت‌ها"),v.get("مسیر تصویر"),_date(v.get("تاریخ انتشار")),_date(v.get("تاریخ پایان")),v.get("متن تماس"),v.get("وضعیت")))
            inserted["special_opportunities"] += 1

        for item in by_name.get("رتبه کارشناسی مناطق", {}).get("rows", []):
            v=item["values"]
            region=conn.execute("SELECT id FROM app.regions WHERE name=%s LIMIT 1", (v.get("منطقه"),)).fetchone()
            conn.execute("INSERT INTO app.region_ranking_entries(period,region_id,rank,label,notes) VALUES (%s,%s,%s,%s,%s)", (v.get("دوره"),region["id"] if region else None,v.get("رتبه ۱ تا ۱۰"),v.get("برچسب"),v.get("توضیحات")))
            inserted["region_rankings"] += 1

        skipped = []
    return {"inserted": inserted, "skipped_sheets": skipped, "message": "داده‌ها با موفقیت ثبت شدند"}
